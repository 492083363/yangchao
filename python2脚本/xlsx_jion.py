#!/env/python
#coding=utf-8

##合并多个excel表格##

import os
import glob
import openpyxl


def merge_xlsx_files(xlsx_files):  #定义函数合并xlsx文件
    wb = openpyxl.load_workbook(xlsx_files[0]) #调用openpyxl模块loda_workbook函数
    ws = wb.active                              #获取活跃的Worksheet
    ws.title = "merged result"                  #定义工作表标题

    for filename in xlsx_files[1:]:            #循环xlsx_files参数，获取第一个工作表（只有一个）
        workbook = openpyxl.load_workbook(filename)  #调用函数
        sheet = workbook.active                   #获取活跃的表格
        for row in sheet.iter_rows(min_row=2): #遍历其他文件，忽略首行内容
            values = [cell.value for cell in row] #循环获取单元格的值
            ws.append(values)                      #将值依次添加末尾    
    return wb                                     #返回

def get_all_xlsx_files(path):                        #定义获取所有xlsx文件
    xlsx_files = glob.glob(os.path.join(path,'*.xlsx')) #采用glob方法指定路径下所有.xlsx的文件
    sorted(xlsx_files,key=str.lower)                     #按照关键字字符串小写排序
    return xlsx_files

def main():                                        #定义主函数
    xlsx_files = get_all_xlsx_files(os.path.expanduser ("E:\\excel\\")) #定义变量xlsx_files为get_all_xlsx_files函数，指定参数为指定目录
    wb = merge_xlsx_files(xlsx_files)                            #定义wb为merge_xlsx_files函数，指定参数为遍历
    wb.save('merged_form.xlsx')                                  #save方法将汇总表保存到merged_form.xlsx

if __name__ =='__main__':
    main()
